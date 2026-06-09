#!/usr/bin/env python3
import copy
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def norm(value):
    return re.sub(r"\s+", "", str(value or "")).strip()


def find_columns(ws):
    best = None
    for row in range(1, min(ws.max_row, 30) + 1):
      values = [norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
      if not any(values):
          continue

      name_cols = [
          index + 1
          for index, value in enumerate(values)
          if "科目" in value and not any(word in value for word in ("编号", "序号", "代码"))
      ]
      code_cols = [
          index + 1
          for index, value in enumerate(values)
          if any(word in value for word in ("科目编号", "编号", "代码", "序号"))
      ]

      if name_cols:
          best = (row, code_cols[0] if code_cols else None, name_cols[0])
          break

    if best:
        return best

    if ws.max_column >= 2:
        return 1, 1, 2
    return 1, None, 1


def clone_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)


def fill_template(template_path, output_path, subjects):
    wb = load_workbook(template_path)
    ws = wb.active
    header_row, code_col, name_col = find_columns(ws)
    first_data_row = header_row + 1

    style_row = first_data_row if first_data_row <= ws.max_row else header_row
    for offset, subject in enumerate(subjects):
        row = first_data_row + offset
        clone_row_style(ws, style_row, row)
        if code_col:
            ws.cell(row, code_col).value = offset + 1
        ws.cell(row, name_col).value = subject

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: fill_subject_template.py <template.xlsx> <output.xlsx> <subjects-json>")

    subjects = json.loads(sys.argv[3])
    subjects = [str(item).strip() for item in subjects if str(item).strip()]
    if not subjects:
        raise SystemExit("No subjects to write")

    fill_template(sys.argv[1], sys.argv[2], subjects)


if __name__ == "__main__":
    main()
